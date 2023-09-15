"""
This script is used to migrate a repository tree and was developed for the HBA
migration.
    bin/instance run ./scripts/repository_migration.py xls_path

optional arguments:
  -o : path to a folder where output should be saved. The folder will be created.
       This defaults to var/migration-TIMESTAMP
  -t : sync tasks. By default tasks are not synced as the migration is not
       performed on the original OGDS. Instead UIDs of tasks that will need
       syncing are stored and dumped in a json file that can then be used
       to sync the tasks.
  -n : dry-run.

If task syncing was skipped, it can be later performed in debug mode:

from opengever.maintenance import dm; dm()
from opengever.maintenance.scripts.repository_migration import TaskSyncer
import json
import transaction
with open(path/to/tasks_to_sync.json, "r") as infile:
    tasks_to_sync = json.load(infile)
TaskSyncer(tasks_to_sync)()
transaction.commit()

Notes:
- permissions for positions that get merged are disregarded
- Setting new permissions will replace the existing sharing permissions.
- Metadata (classification, privacy_layer, retention_period,
  retention_period_annotation, archival_value, archival_value_annotation,
  custody_period) are only handled for new object creations.
"""

from Acquisition import aq_inner
from Acquisition import aq_parent
from collections import defaultdict
from collections import namedtuple
from collective.transmogrifier.transmogrifier import Transmogrifier
from copy import deepcopy
from opengever.base.behaviors.classification import IClassification
from opengever.base.behaviors.lifecycle import ILifeCycle
from opengever.base.default_values import get_persisted_value_for_field
from opengever.base.default_values import object_has_value_for_field
from opengever.base.indexes import sortable_title
from opengever.base.interfaces import IReferenceNumber
from opengever.base.interfaces import IReferenceNumberFormatter
from opengever.base.interfaces import IReferenceNumberPrefix
from opengever.base.monkey.patching import MonkeyPatch
from opengever.base.role_assignments import ASSIGNMENT_VIA_SHARING
from opengever.base.role_assignments import RoleAssignmentManager
from opengever.base.role_assignments import SharingRoleAssignment
from opengever.base.schemadump.config import ROLES_BY_SHORTNAME
from opengever.base.utils import unrestrictedUuidToObject
from opengever.bundle.console import add_guid_index
from opengever.bundle.ldap import DisabledLDAP
from opengever.bundle.sections.bundlesource import BUNDLE_PATH_KEY
from opengever.bundle.sections.commit import INTERMEDIATE_COMMITS_KEY
from opengever.bundle.sections.constructor import BUNDLE_GUID_KEY
from opengever.dossier.behaviors.dossier import IDossierMarker
from opengever.globalindex.handlers.task import TaskSqlSyncer
from opengever.maintenance.debughelpers import setup_app
from opengever.maintenance.debughelpers import setup_option_parser
from opengever.maintenance.debughelpers import setup_plone
from opengever.maintenance.scripts.update_object_ids import ObjectIDUpdater
from opengever.repository.behaviors import referenceprefix
from opengever.repository.deleter import RepositoryDeleter
from opengever.repository.interfaces import IRepositoryFolder
from opengever.repository.interfaces import IRepositoryFolderRecords
from opengever.setup.sections.xlssource import MAPPED_FIELDS
from opengever.setup.sections.xlssource import xlrd_xls2array
from openpyxl import Workbook
from openpyxl.styles import Font
from plone import api
from plone.app.uuid.utils import uuidToCatalogBrain
from plone.uuid.interfaces import IUUID
from Products.CMFPlone.utils import safe_unicode
from uuid import uuid4
from zope.annotation import IAnnotations
from zope.component import queryAdapter
import json
import logging
import os
import shutil
import sys
import tempfile
import time
import transaction

logger = logging.getLogger('opengever.maintenance')
handler = logging.StreamHandler(stream=sys.stdout)
logging.root.addHandler(handler)
logging.root.setLevel(logging.INFO)


MIGRATION_KEY = 'opengever.maintenance.repository_migration'
MIGRATIOM_TIMESTAMP = time.strftime('%d%m%Y-%H%M%S')
tasks_to_sync = set()

managed_roles_shortnames = ['read', 'add', 'edit', 'close', 'reactivate', 'manage_dossiers']
SHORTNAMES_BY_ROLE = {value: key for key, value in ROLES_BY_SHORTNAME.items()}

metadata_fields = (
    IClassification["classification"],
    IClassification["privacy_layer"],
    ILifeCycle["retention_period"],
    ILifeCycle["retention_period_annotation"],
    ILifeCycle["archival_value"],
    ILifeCycle["archival_value_annotation"],
    ILifeCycle["custody_period"],
                   )

MAPPED_FIELDS = deepcopy(MAPPED_FIELDS)
# Mapping used in StaSG migration for some reason
MAPPED_FIELDS["archival_value"].update({
    u'ng': u'unchecked',
    u'an': u'prompt',
    u'ar': u'archival worthy',
    u'na': u'not archival worthy',
    u'sa': u'archival worthy with sampling',
                                       })


def log_progress(i, tot, step=100):
    if i % step == 0:
        logger.info(u'{}: Done {} / {}'.format(
            time.strftime('%d.%m.%Y %H:%M:%S'), i, tot))


class MigrationPreconditionsError(Exception):
    """Raised when errors are found during migration validation"""


class MigrationValidationError(Exception):
    """Raised when errors are found during migration validation"""


class PatchCommitSection(MonkeyPatch):
    """To maintain transactionality we need to avoid that the bundle import
    commits changes made. Here we patch out commits from the CommitSection.
    """

    def __call__(self):
        from opengever.bundle.sections.commit import CommitSection

        def __iter__(self):
            for count, item in enumerate(self.previous, start=1):
                if count % self.every == 0 and self.intermediate_commits:
                    logger.info("skipping intermediate commit after %s items..." % count)

                yield item

            logger.info("Skipping commit after bundle import...")

        self.patch_refs(CommitSection, '__iter__', __iter__)


class PatchReindexContainersSection(MonkeyPatch):
    """To maintain transactionality we need to avoid that the bundle import
    commits changes made. Here we patch out commits from the ReindexContainersSection.
    """

    def __call__(self):
        from opengever.bundle.sections.reindex_containers import ReindexContainersSection
        from collective.transmogrifier.utils import traverse

        def __iter__(self):
            for item in self.previous:
                yield item

            n_containers = len(self.bundle.containers_to_reindex)
            logger.info("Reindexing {} containers after bundle import...".format(n_containers))

            for container_path in self.bundle.containers_to_reindex:
                obj = traverse(self.site, container_path, None)
                obj.reindexObject(idxs=self.indexes)

            logger.info("Skipping commit...")

        self.patch_refs(ReindexContainersSection, '__iter__', __iter__)


class PatchReportSection(MonkeyPatch):
    """To maintain transactionality we need to avoid that the bundle import
    commits changes made. As we also do not need the reports for a migration,
    we patch out the whole ReportSection.
    """

    def __call__(self):
        from opengever.bundle.sections.report import ReportSection

        def __iter__(self):
            for item in self.previous:
                yield item

        self.patch_refs(ReportSection, '__iter__', __iter__)


class PatchDisableLDAP(MonkeyPatch):
    """To maintain transactionality we need to avoid that the bundle import
    commits changes made. Here we patch out commits from the DisableLDAP
    context manager.
    """

    def __call__(self):
        from opengever.bundle.ldap import DisabledLDAP
        from opengever.bundle.ldap import enable_ldap

        def __exit__(self, exc_type, exc_val, exc_tb):
            if exc_val is not None:
                # Exception happened, make sure transaction is rolled back
                transaction.abort()
                transaction.begin()

            enable_ldap(self.portal)

            # Make sure persistent changes that re-enable LDAP are committed
            # transaction.commit()

        self.patch_refs(DisabledLDAP, '__exit__', __exit__)


class SkipTaskSyncWith(MonkeyPatch):
    """ We skip syncing the tasks altogether, as migration is not done with the
    productive OGDS. We will then sync the tasks at a later stage. We therefore store
    the UIDs of tasks for later use.
    """

    def __call__(self):
        from opengever.globalindex.model.task import Task

        def sync_with(self, plone_task, graceful=False):
            """Sync this task instace with its corresponding plone taks."""
            tasks_to_sync.add(plone_task.UID())
            return
        self.patch_refs(Task, 'sync_with', sync_with)


class SkipDocPropsUpdate(MonkeyPatch):
    """ No nead to update the docproperties, we anyway don't have the
    blobs during the migration
    """

    def __call__(self):
        from opengever.document import handlers

        def _update_docproperties(document, raise_on_error=False):
            return

        self.patch_refs(handlers, '_update_docproperties', _update_docproperties)


class SkipSearchableTextExtraction(MonkeyPatch):
    """ During migration we do not have the blobs, so that we should
    avoid extracting full text from the blobs.
    """

    def __call__(self):
        from ftw.solr.connection import SolrConnection

        def extract(self, blob, field, data, content_type):
            return

        self.patch_refs(SolrConnection, 'extract', extract)


def cleanup_position(position):
    """Remove splitting dots - they're not usefull for comparison.
    This only works for grouped_by_three formatter.
    """
    if position is None:
        return None
    position = str(position).strip()
    if position:
        return position.replace('.', '')


class RepositoryPosition(object):

    def __init__(self, position=None, title=None, description=None):
        self.position = cleanup_position(position)
        self.title = self.to_safe_unicode(title)
        self.description = self.to_safe_unicode(description)

    @staticmethod
    def to_safe_unicode(maybe_none):
        if maybe_none is None:
            return None
        maybe_none = safe_unicode(maybe_none)
        if maybe_none:
            return maybe_none

    @property
    def reference_number_prefix(self):
        """Returns last part of the position - the referencenumber prefix"""
        return self.position[-1]

    @property
    def parent_position(self):
        """Returns the position without the last part of the position, i.e.
        the parent position"""
        return self.position[:-1]

    def __repr__(self):
        return u"RepositoryPosition({}, {}, {})".format(self.position, self.title, self.description).encode('utf-8')

    def __eq__(self, other):
        return all((self.position == other.position,
                    self.title == other.title,
                    self.description == other.description))


class Row(object):

    def __init__(self, row, column_mapping):
        for key, column in column_mapping.items():
            col = column.index
            setattr(self, key, row[col])


Column = namedtuple('Column', ('index', 'technical_header', 'header'))


class ExcelDataExtractor(object):

    header_row = 2
    technical_header_row = 4
    first_data_row = 6

    column_mapping = {
        'old_position': Column(0, '', u'Ordnungs-\npositions-\nnummer'),  # A
        'old_title': Column(1, '', u'Titel der Ordnungsposition'),  # B
        'old_description': Column(2, '', u'Beschreibung (optional)'),  # C
        'new_position': Column(5, 'reference_number',
                               u'Ordnungs-\npositions-\nnummer'),  # F
        'new_title': Column(6, 'effective_title', u'Titel der Ordnungsposition'),  # G
        'new_description': Column(8, 'description', u'Beschreibung (optional)'),  # I
        'classification': Column(11, 'classification', u'Klassifikation'),  # L
        'privacy_layer': Column(12, 'privacy_layer', u'Datenschutz'),  # M
        'retention_period': Column(13, 'retention_period',
                                   u'Aufbewahrung in Verwaltung'),  # N
        'retention_period_annotation': Column(14, 'retention_period_annotation',
                                              u'Kommentar zur Aufbewahrungsdauer\n(optional)'),  # O
        'archival_value': Column(15, 'archival_value', u'Archiv-\nw\xfcrdigkeit'),  # P
        'archival_value_annotation': Column(16, 'archival_value_annotation',
                                            u'Kommentar zur Archivw\xfcrdigkeit\n(optional)'),  # Q
        'custody_period': Column(17, 'custody_period', u'Archivische Schutzfrist'),  # R
        'block_inheritance': Column(22, 'block_inheritance', ''),  # W
        'read': Column(23, 'read_dossiers_access', ''),  # X
        'add': Column(24, 'add_dossiers_access', ''),  # Y
        'edit': Column(25, 'edit_dossiers_access', ''),  # Z
        'close': Column(26, 'close_dossiers_access', ''),  # AA
        'reactivate': Column(27, 'reactivate_dossiers_access', ''),  # AB
        'manage_dossiers': Column(28, 'manage_dossiers_access', ''),  # AC
    }

    def __init__(self, diff_xlsx_path):
        sheets = xlrd_xls2array(diff_xlsx_path)
        self.data = sheets[0]['sheet_data']
        self.n_data = len(self.data) - self.first_data_row
        self.validate_format()

    def validate_format(self):
        headers = self.data[self.header_row]
        technical_headers = self.data[self.technical_header_row]
        for column in self.column_mapping.values():
            assert technical_headers[column.index] == column.technical_header, \
                u"Column technical header mismatch: {} != {}".format(
                    technical_headers[column.index], column.technical_header)

            if not column.header:
                continue

            assert headers[column.index] == column.header, \
                u"Column header mismatch: {} != {}".format(
                    headers[column.index], column.header)

    def get_data(self):
        for row in self.data[self.first_data_row:]:
            yield Row(row, self.column_mapping)


class PositionsMapping(object):
    """Mapping of both existing and newly created positions.
    Maps old_position_numbers and new_position_numbers to
    obj and guid.
    """

    def __init__(self, operations, reference_repository_mapping):
        logger.info(u"\n\nPreparing new positions mapping...\n")
        self.old_pos_guid = {}
        self.new_pos_guid = {}
        self.old_pos_new_guid = {}
        self.reference_repository_mapping = reference_repository_mapping
        self._create_mapping(operations)

    def _add_creation(self, operation):
        new_refnum = operation['new_repo_pos'].position
        if new_refnum in self.new_pos_guid:
            # we have a creation operation for a position that already
            # exists and does not change, this should not happen
            raise Exception("Useless creation operation for {}".format(new_refnum))
        self.new_pos_guid[new_refnum] = uuid4().hex[:8]

    def _add_move(self, operation):
        old_refnum = operation['old_repo_pos'].position
        new_refnum = operation['new_repo_pos'].position

        if old_refnum in self.old_pos_guid:
            # only one row allowed for each existing position
            logger.warning(
                "\nInvalid operation: position appears twice in excel."
                " {}\n".format(operation))
            operation['is_valid'] = False
        if not new_refnum:
            raise Exception("This should not happen, we normally skip deletions")

        obj = self.get_object_for_position(old_refnum)
        guid = IAnnotations(obj)[BUNDLE_GUID_KEY]

        self.old_pos_guid[old_refnum] = guid
        if new_refnum not in self.new_pos_guid:
            self.new_pos_guid[new_refnum] = guid
        else:
            # Will get merged, let's remember into which guid
            self.old_pos_new_guid[old_refnum] = self.new_pos_guid[new_refnum]

    def get_old_pos_guid(self, old_refnum):
        """Maps an old position to its guid"""
        return self.old_pos_guid.get(old_refnum)

    def get_old_pos_new_guid(self, old_refnum):
        """Maps an old position to its new guid, only if it changes (merges)"""
        return self.old_pos_new_guid.get(old_refnum)

    def get_new_pos_guid(self, new_refnum):
        """Maps a new position to its guid"""
        return self.new_pos_guid.get(new_refnum)

    def get_object_for_position(self, position):
        return self.reference_repository_mapping.get(position)

    def _create_mapping(self, operations):
        """
        we split the operations so as to first treat all rows where the reference
        number is not changed, then creation operations and finally move and
        merge operations. This will help optimize the operations we will
        execute in the end, as new_pos_guid will contain the guid of
        the object corresponding from the first operation pointing to that
        position, all other operations pointing to that new position will be
        considered merges. We therefore want to make sure that if one position
        is unchanged and another merged into it, we do not instead move
        the second one and merge the first one into it instead (same result
        but inefficient).
        """
        unchanged = []
        creations = []
        moves_or_merges = []
        for operation in operations:
            old_repo_pos = operation['old_repo_pos']
            new_repo_pos = operation['new_repo_pos']
            if old_repo_pos.position == new_repo_pos.position:
                unchanged.append(operation)
            elif old_repo_pos.position is None:
                creations.append(operation)
            else:
                moves_or_merges.append(operation)

        for operation in unchanged:
            self._add_move(operation)

        for operation in creations:
            self._add_creation(operation)

        for operation in moves_or_merges:
            self._add_move(operation)

    def is_complete(self):
        """Now we make sure that the excel was complete, i.e. there was a row
        for each existing repository folder"""
        complete = True
        for ref_num in self.reference_repository_mapping:
            if ref_num not in self.old_pos_guid:
                logger.warning("\nExcel is incomplete. No operation defined for "
                               "position {}\n".format(ref_num))
                complete = False
        return complete


class MigratorBase(object):

    def guid_to_object(self, guid):
        results = self.catalog.unrestrictedSearchResults(bundle_guid=guid)
        if len(results) == 0:
            # This should never happen. Object with a guid should have been created.
            logger.warning(
                u"Couldn't find object with GUID %s in catalog" % guid)
            return

        if len(results) > 1:
            # Ambiguous GUID - this should never happen
            logger.warning(
                u"Ambiguous GUID! Found more than one result in catalog "
                u"for GUID %s " % guid)
            return

        return results[0].getObject()


class RepositoryExcelAnalyser(MigratorBase):

    def __init__(self, mapping_path, output_directory):
        self.number_changes = {}

        self.diff_xlsx_path = mapping_path
        self.output_directory = output_directory
        self.analysed_rows = []
        self._reference_repository_mapping = None
        self.catalog = api.portal.get_tool('portal_catalog')
        self.is_valid = True

        self.check_preconditions()
        self.prepare_guids()
        self.reporoot, self.reporoot_guid = self.get_reporoot_and_guid()

        self.new_positions = set()
        self.logged_metadata_mismatch = False

    def check_preconditions(self):
        logger.info(u"\n\nChecking preconditions...\n")
        # current implementation only works with grouped_by_three reference
        # formatter, notably because we remove splitting dots during the analysis.
        formatter_name = api.portal.get_registry_record(
            "opengever.base.interfaces.IReferenceNumberSettings.formatter")
        assert formatter_name == "grouped_by_three", "Migration is only supported with grouped_by_three"
        self.formatter = queryAdapter(
                api.portal.get(), IReferenceNumberFormatter, name=formatter_name)

        # Creation of new repository folders in the repository root will only
        # work if their is a single repository root.
        results = self.catalog.unrestrictedSearchResults(
            portal_type='opengever.repository.repositoryroot')
        assert len(results) == 1, "Migration is only supported with a single repository root"

    def prepare_guids(self):
        """ The GUID index is needed by the bundle transmogrifier.
        Moreover the repository root needs to have a GUID, as it does not
        have a reference number allowing to find the parent when creating
        a new repository folder in the repository root.
        We also add a guid for all repository folders, as these allow to
        unequivocally identify a parent when creating a new repofolder.
        """
        logger.info(u"\n\nPreparing GUIDs...\n")
        add_guid_index()
        brains = self.catalog.unrestrictedSearchResults(
            portal_type=['opengever.repository.repositoryroot',
                         'opengever.repository.repositoryfolder']
            )
        for brain in brains:
            obj = brain.getObject()
            if not IAnnotations(obj).get(BUNDLE_GUID_KEY):
                IAnnotations(obj)[BUNDLE_GUID_KEY] = uuid4().hex[:8]
            # reindex, whether the guid was already here or not, to make sure
            # it's also in the catalog
            obj.reindexObject(idxs=['bundle_guid'])

    def get_reporoot_and_guid(self):
        brains = self.catalog.unrestrictedSearchResults(
            portal_type='opengever.repository.repositoryroot')
        reporoot = brains[0].getObject()
        return reporoot, IAnnotations(reporoot)[BUNDLE_GUID_KEY]

    def extract_data(self):
        logger.info(u"\n\nExtracting data from Excel...\n")
        data_extractor = ExcelDataExtractor(self.diff_xlsx_path)

        operations = []
        for i, row in enumerate(data_extractor.get_data()):
            log_progress(i, data_extractor.n_data)
            if row.new_position in ['', u'l\xf6schen', '-']:
                # Position should be deleted
                new_repo_pos = RepositoryPosition()
            else:
                new_repo_pos = RepositoryPosition(row.new_position, row.new_title, row.new_description)
            if row.old_position == '':
                # Position did not exist
                old_repo_pos = RepositoryPosition()
            else:
                old_repo_pos = RepositoryPosition(row.old_position, row.old_title, row.old_description)

            # Ignore empty rows
            if not old_repo_pos.position and not new_repo_pos.position:
                continue

            # Skip positions that should be deleted
            if not new_repo_pos.position:
                logger.info("\nSkipping, we do not support deletion: {}\n".format(row))
                continue

            permissions = self.extract_permissions(row)
            metadata = self.extract_metadata(row)
            operations.append({
                'old_repo_pos': old_repo_pos,
                'new_repo_pos': new_repo_pos,
                'permissions': permissions,
                'metadata': metadata,
                'is_valid': True})
        return operations

    def analyse(self):
        logger.info(u"\n\nStarting analysis...\n")
        data = self.extract_data()

        self.positions_mapping = PositionsMapping(data, self.get_repository_reference_mapping())
        # add the repository root to the mapping
        self.positions_mapping.old_pos_guid[''] = self.reporoot_guid
        self.positions_mapping.new_pos_guid[''] = self.reporoot_guid
        if not self.positions_mapping.is_complete():
            self.is_valid = False

        for operation in data:
            old_repo_pos = operation['old_repo_pos']
            new_repo_pos = operation['new_repo_pos']

            new_number = None
            new_parent_guid = None
            new_position_guid = None
            old_parent_guid = None

            need_creation = not bool(old_repo_pos.position)
            need_number_change, need_move, need_merge = self.needs_number_change_move_or_merge(operation)

            if need_number_change:
                new_number = self.get_new_number(new_repo_pos)
                old_parent_guid = self.positions_mapping.get_old_pos_guid(old_repo_pos.parent_position)
            if need_move:
                new_parent_guid = self.get_new_parent_guid(new_repo_pos)
                old_parent_guid = self.positions_mapping.get_old_pos_guid(old_repo_pos.parent_position)
            if need_merge:
                new_parent_guid = self.positions_mapping.get_new_pos_guid(new_repo_pos.position)
                old_parent_guid = self.positions_mapping.get_old_pos_guid(old_repo_pos.parent_position)
            if need_creation:
                new_parent_guid = self.get_new_parent_guid(new_repo_pos)
                new_position_guid = self.positions_mapping.get_new_pos_guid(new_repo_pos.position)

            operation.update({
                'uid': self.get_uuid_for_position(old_repo_pos.position),
                'new_position_guid': new_position_guid,
                'need_move': need_move,
                'need_merge': need_merge,
                'need_creation': need_creation,
                'new_title': self.get_new_title(new_repo_pos, old_repo_pos) if not (need_creation or need_merge) else None,
                'new_number': new_number,
                'new_parent_guid': new_parent_guid,
                'old_parent_guid': old_parent_guid})

            self.validate_operation(operation)

            self.analysed_rows.append(operation)

        # Make sure that analysis is invalid if any operation was invalid
        if any([not op['is_valid'] for op in self.analysed_rows]):
            self.is_valid = False

    def validate_operation(self, operation):
        """Make sure that operation satisfies all necessary conditions and add
        is_valid, repository_depth_violated and leaf_node_violated and
        permissions_disregarded to it.
        """
        # Each operation should either have a uid or a new_position_guid
        if not any((operation['new_position_guid'], operation['uid'])):
            logger.warning("\nInvalid operation: needs new_position_guid "
                           "or uid. {}\n".format(operation))
            operation['is_valid'] = False

        # Make sure that all UIDs are valid and that for existing UIDs,
        # the title, position and description match the ones in the Excel.
        # We also check that the provided metadata match with the existing
        # object's but only issue a warning if they don't.
        operation['metadata_mismatch'] = False
        if operation['uid']:
            obj = unrestrictedUuidToObject(operation['uid'])
            if not obj:
                logger.warning("\nInvalid operation: uid is not valid."
                               "or uid. {}\n".format(operation))
                operation['is_valid'] = False

            else:
                old_repo_pos = operation['old_repo_pos']
                if obj.title_de != old_repo_pos.title:
                    logger.warning("\nInvalid operation: incorrect title."
                                   "{}\n".format(operation))
                    operation['is_valid'] = False
                if obj.get_repository_number().replace('.', '') != old_repo_pos.position:
                    logger.warning("\nInvalid operation: incorrect position."
                                   "{}\n".format(operation))
                    operation['is_valid'] = False
                if (obj.description or old_repo_pos.description) and (obj.description or '').strip() != (old_repo_pos.description or '').strip():
                    logger.warning("\nInvalid operation: incorrect description."
                                   "{}\n".format(operation))
                    operation['is_valid'] = False

                self.check_metadata(operation, obj)

        # Each operation should have new position
        if not operation['new_repo_pos'].position:
            logger.warning("\nInvalid operation: needs new position. {}\n".format(
                operation))
            operation['is_valid'] = False

        if all((operation['new_position_guid'], operation['uid'])):
            logger.warning("\nInvalid operation: can define only one of "
                           "new_position_guid or uid. {}\n".format(operation))
            operation['is_valid'] = False

        # A move operation should have a new_parent_guid
        if operation['need_move']:
            if not operation['new_parent_guid']:
                logger.warning(
                    "\nInvalid operation: move operation must define "
                    "new_parent_guid. {}\n".format(operation))
                operation['is_valid'] = False

        # Make sure that if a position is being created, its parent will be found
        if operation['need_creation']:
            if not operation['new_parent_guid'] in self.positions_mapping.new_pos_guid.values():
                logger.warning(
                    "\nInvalid operation: could not find new parent for create "
                    "operation. {}\n".format(operation))
                operation['is_valid'] = False

        self.check_repository_depth_violation(operation)
        self.check_leaf_node_principle_violation(operation)

        # Each new position can only have one row in the excel except for merge operations
        new_position = operation['new_repo_pos'].position
        if new_position and not operation['need_merge']:
            if new_position in self.new_positions:
                logger.warning(
                    "\nInvalid operation: new position appears twice in excel."
                    " {}\n".format(operation))
                operation['is_valid'] = False
            self.new_positions.add(new_position)

        # if position is being merged, then permissions set in that row will
        # be lost. Best would be to compare the permissions of that row with
        # the ones it gets merged into. Instead we simply log and write it
        # in the analysis excel. The user can make sure this is correct himself.
        permissions = operation['permissions']
        operation['permissions_disregarded'] = False
        operation['local_roles_deleted'] = False
        operation['set_permissions'] = False
        if operation['need_merge']:
            if any(permissions.values()):
                logger.info(
                    "\nPermissions disregarded: this position gets merged"
                    " {}\n".format(operation))
                operation['permissions_disregarded'] = True
        else:
            # We make sure that when inheritance is blocked, local roles are set
            has_local_roles = any(permissions[role_shortname] for role_shortname in managed_roles_shortnames)
            inheritance_blocked = permissions['block_inheritance']
            if inheritance_blocked and not has_local_roles:
                logger.warning(
                    "\nInvalid operation: blocking inheritance without setting "
                    "local roles. {}\n".format(operation))
                operation['is_valid'] = False
            obj = unrestrictedUuidToObject(operation['uid'])
            if not obj:
                # newly created positions will have the local_roles set
                # in the pipeline
                pass
            elif self.needs_permission_update(obj, permissions, inheritance_blocked):
                operation['set_permissions'] = True
                if RoleAssignmentManager(obj).get_assignments_by_cause(ASSIGNMENT_VIA_SHARING):
                    operation['local_roles_deleted'] = True
                    logger.warning(
                        "\nSharing assignments for {} will be deleted and "
                        "replaced.\n".format(obj.absolute_url_path()))

    def needs_permission_update(self, obj, permissions, inheritance_blocked):
        if inheritance_blocked != getattr(obj, '__ac_local_roles_block__', False):
            return True
        old_permissions = {role_shortname: [] for role_shortname in managed_roles_shortnames}
        assignments = RoleAssignmentManager(obj).get_assignments_by_cause(ASSIGNMENT_VIA_SHARING)
        for assignment in assignments:
            for role in assignment['roles']:
                if role not in SHORTNAMES_BY_ROLE or SHORTNAMES_BY_ROLE[role] not in managed_roles_shortnames:
                    logger.info("Skipping {} role for {} when checking for permissions update for {}.".format(
                        role, assignment['principal'], obj.absolute_url()))
                    continue
                shortname = SHORTNAMES_BY_ROLE[role]
                old_permissions[shortname].append(assignment['principal'])

        for role_shortname in managed_roles_shortnames:
            if not set(old_permissions[role_shortname]) == set(permissions[role_shortname]):
                return True
        return False

    def get_new_title(self, new_repo_pos, old_repo_pos):
        """Returns the new title or none if no rename is necessary."""
        if new_repo_pos.title != old_repo_pos.title:
            return new_repo_pos.title

        return None

    def get_new_number(self, new_repo_pos):
        """Returns latest part of the position - the new referencenumber
        prefix"""
        return new_repo_pos.reference_number_prefix

    def get_new_parent_guid(self, new_repo_pos):
        """Returns the new parent guid."""
        return self.positions_mapping.get_new_pos_guid(new_repo_pos.parent_position)

    def needs_number_change_move_or_merge(self, operation):
        """Check if a number change, a move or a merge is necessary
        """
        need_number_change = False
        need_move = False
        need_merge = False

        old_repo_pos = operation['old_repo_pos']
        new_repo_pos = operation['new_repo_pos']
        old_refnum = old_repo_pos.position
        new_refnum = new_repo_pos.position

        if not (new_refnum and old_refnum):
            # creation or deletion
            return need_number_change, need_move, need_merge

        # guid change is a merge operation
        if self.positions_mapping.get_old_pos_new_guid(old_refnum):
            need_merge = True
            return need_number_change, need_move, need_merge

        old_parent_pos_guid = self.positions_mapping.get_old_pos_guid(old_repo_pos.parent_position)
        new_parent_pos_guid = self.positions_mapping.get_new_pos_guid(new_repo_pos.parent_position)
        if not new_parent_pos_guid:
            logger.warning(
                "\nInvalid operation: cannot find new parent. "
                "{}\n".format(operation))
            operation['is_valid'] = False
            return need_number_change, need_move, need_merge

        # move operation is when parent changes except if the parent is
        # merged into the new parent
        if old_parent_pos_guid != new_parent_pos_guid:
            old_parent_new_guid = self.positions_mapping.get_old_pos_new_guid(old_repo_pos.parent_position)
            # if current parent is merged into the future parent, no need to move
            if old_parent_new_guid != new_parent_pos_guid:
                need_move = True

        # check if number change is necessary
        if need_move or new_repo_pos.reference_number_prefix != old_repo_pos.reference_number_prefix:
            need_number_change = True

        return need_number_change, need_move, need_merge

    def check_repository_depth_violation(self, operation):
        max_depth = api.portal.get_registry_record(
            interface=IRepositoryFolderRecords, name='maximum_repository_depth')

        new_repo_pos = operation['new_repo_pos']
        if new_repo_pos.position and len(new_repo_pos.position) > max_depth:
            logger.warning("\nInvalid operation: repository depth violated."
                           " {}\n".format(operation))
            operation['is_valid'] = False
            operation['repository_depth_violated'] = True
        else:
            operation['repository_depth_violated'] = False

    def check_leaf_node_principle_violation(self, operation):
        operation['leaf_node_violated'] = False
        if not (operation['need_move'] or operation['need_creation']):
            # object is neither moved nor created, nothing to worry about
            return

        if operation['new_parent_guid'] not in self.positions_mapping.old_pos_guid.values():
            # parent is being created, hard to check leaf node principle
            return
        parent_repo = self.guid_to_object(operation['new_parent_guid'])

        if not parent_repo:
            # Something is fishy, parent should either exist or be created
            operation['is_valid'] = False
            logger.warning("\nInvalid operation: parent not found. {}\n".format(operation))
            return
        if any([IDossierMarker.providedBy(item) for item in parent_repo.objectValues()]):
            operation['is_valid'] = False
            operation['leaf_node_violated'] = True
            logger.warning("\nInvalid operation: leaf node principle violated."
                           " {}\n".format(operation))

    def get_repository_reference_mapping(self):
        if not self._reference_repository_mapping:
            repos = [brain.getObject() for brain in
                     self.catalog(object_provides=IRepositoryFolder.__identifier__)]
            self._reference_repository_mapping = {
                repo.get_repository_number().replace('.', ''): repo for repo in repos}

        return self._reference_repository_mapping

    def get_uuid_for_position(self, position):
        mapping = self.get_repository_reference_mapping()

        if position and position in mapping:
            return IUUID(mapping[position])

        return None

    def extract_permissions(self, row):
        permissions = {'block_inheritance': False}

        if row.block_inheritance:
            block = row.block_inheritance.strip()
            assert block in ['ja', 'nein']
            if block == 'ja':
                permissions['block_inheritance'] = True

        for key in managed_roles_shortnames:
            groups = [group.strip() for group in getattr(row, key).split(',')]
            groups = [group for group in groups if group]

            permissions[key] = groups

        return permissions

    def extract_metadata(self, row):
        metadata = {}
        for field in metadata_fields:
            fieldname = field.getName()
            value = getattr(row, fieldname)

            if value in (None, '', u''):
                continue

            if fieldname in MAPPED_FIELDS.keys():
                mapping = MAPPED_FIELDS[fieldname]

                # Data not already a valid term
                if value not in mapping.values():
                    value = mapping[value.lower()]

            metadata[fieldname] = value
        return metadata

    def check_metadata(self, operation, obj):
        metadata = operation["metadata"]
        try:
            for field in metadata_fields:
                if object_has_value_for_field(obj, field):
                    persisted_value = get_persisted_value_for_field(obj, field)
                else:
                    persisted_value = None
                assert(persisted_value == metadata.get(field.getName()))
        except AssertionError:
            operation['metadata_mismatch'] = True
            if not self.logged_metadata_mismatch:
                logger.warning("\nMetadata mismatch. At least one operation has"
                               " a metadata mismatch. Metadata will not be "
                               "updated during migration!")
                self.logged_metadata_mismatch = True

    def export_to_excel(self):
        analyse_xlsx_path = os.path.join(self.output_directory, 'analysis.xlsx')
        workbook = self.prepare_workbook(self.analysed_rows)
        # Save the Workbook-data in to a StringIO
        return workbook.save(filename=analyse_xlsx_path)

    def prepare_workbook(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Analyse'

        self.insert_label_row(sheet)
        self.insert_value_rows(sheet, rows)

        return workbook

    def insert_label_row(self, sheet):
        title_font = Font(bold=True)
        labels = [
            # metadata
            'UID',
            'Neu: Position', 'Neu: Titel', 'Neu: Description',
            'Alt: Position', 'Alt: Titel', 'Alt: Description',

            # operations
            'Position Erstellen',
            'Umbenennung (Neuer Titel)',
            'Nummer Anpassung (Neuer `Praefix`)',
            'Verschiebung noetig',
            'Merge noetig',

            # rule violations
            'Verletzt Max. Tiefe',
            'Verletzt Leafnode Prinzip',
            'Ist ungultig',

            # metadata
            'Metadaten stimmen nicht ueberein',

            # permission
            'Ignorierte Bewilligungen',
            'Vorherige Lokalen Rollen entfernt',
            'Bewilligungen',
        ]

        for i, label in enumerate(labels, 1):
            cell = sheet.cell(row=1 + 1, column=i)
            cell.value = label
            cell.font = title_font

    def insert_value_rows(self, sheet, rows):
        for row, data in enumerate(rows, 2):
            values = [
                data['uid'],
                data['new_repo_pos'].position,
                data['new_repo_pos'].title,
                data['new_repo_pos'].description,
                data['old_repo_pos'].position,
                data['old_repo_pos'].title,
                data['old_repo_pos'].description,
                data['need_creation'],
                data['new_title'],
                data['new_number'],
                data['need_move'],
                data['need_merge'],
                'x' if data['repository_depth_violated'] else '',
                'x' if data['leaf_node_violated'] else '',
                'x' if not data['is_valid'] else '',
                'x' if data['metadata_mismatch'] else '',
                'x' if data['permissions_disregarded'] else '',
                'x' if data['local_roles_deleted'] else '',
                json.dumps(data['permissions']) if any(data['permissions'].values()) else '',
            ]

            for column, attr in enumerate(values, 1):
                cell = sheet.cell(row=1 + row, column=column)
                cell.value = attr


class RepositoryMigrator(MigratorBase):

    def __init__(self, operations_list, dry_run=False):
        self.operations_list = operations_list
        self.dry_run = dry_run
        self._reference_repository_mapping = None
        self.to_reindex = defaultdict(set)
        self.catalog = api.portal.get_tool('portal_catalog')
        self.check_preconditions()

    def check_preconditions(self):
        if any(not operation['is_valid'] for operation in self.operations_list):
            raise MigrationPreconditionsError("Some operations are invalid.")

    def run(self):
        self.set_permissions(self.items_to_set_permissions())
        self.create_repository_folders(self.items_to_create())
        self.move_branches(self.items_to_move())
        self.merge_branches(self.items_to_merge())
        self.adjust_reference_number_prefix(self.items_to_adjust_number())
        self.regenerate_reference_number_mapping(self.objects_to_fix_refnum_mapping())
        self.rename(self.items_to_rename())
        self.update_description(self.operations_list)
        self.reindex()
        self.validate()

    def objects_to_fix_refnum_mapping(self):
        """Items that get created over the bundle import will get the
        default value as the reference number prefix is their number is already
        taken, which is then saved in the mapping in saveReferenceNumberPrefix.
        The reference number will then get updated later in the DexterityUpdateSection
        when setting the values for the metadata. This leads to the position
        appearing twice in the mapping on the parent (default reference number
        marked as inactive).
        Also moving repository folders is normally not allowed in Gever and
        it seems the mappings of both the old and new parent objects do not get
        updated correctly. So we need to update them by hand as well.
        """
        parent_guids = set()
        for item in self.items_to_create():
            parent_guids.add(item['new_parent_guid'])

        for item in self.items_to_move():
            parent_guids.add(item['new_parent_guid'])
            parent_guids.add(item['old_parent_guid'])

        for guid in parent_guids:
            obj = self.guid_to_object(guid)
            if not obj:
                # Some objects have been deleted because of merge operations
                continue
            yield obj

    def items_to_create(self):
        return [item for item in self.operations_list if item['new_position_guid']]

    def items_to_move(self):
        return [item for item in self.operations_list if item['need_move']]

    def items_to_merge(self):
        return [item for item in self.operations_list if item['need_merge']]

    def items_to_adjust_number(self):
        return [item for item in self.operations_list if item['new_number']]

    def items_to_rename(self):
        return [item for item in self.operations_list if item['new_title']]

    def items_to_set_permissions(self):
        return [item for item in self.operations_list if item['set_permissions']]

    def add_to_reindexing_queue(self, uid, idxs, with_children=False):
        self.to_reindex[uid].update(idxs)
        obj = unrestrictedUuidToObject(uid)
        if not with_children:
            return

        contained_brains = self.catalog.unrestrictedSearchResults(
            path=obj.absolute_url_path())
        for brain in contained_brains:
            self.to_reindex[brain.UID].update(idxs)

    def create_repository_folders(self, items):
        """Add repository folders - by using the ogg.bundle import. """
        logger.info("\n\nCreating bundle...\n")
        bundle_items = []
        for item in items:
            bundle_item = {
                'guid': item['new_position_guid'],
                'description': item['new_repo_pos'].description,
                'parent_guid': item['new_parent_guid'],
                'reference_number_prefix': item['new_repo_pos'].reference_number_prefix,
                'review_state': 'repositoryfolder-state-active',
                'title_de': item['new_repo_pos'].title,
                '_permissions': item['permissions']
                }
            bundle_item.update(item["metadata"])
            bundle_items.append(bundle_item)

        tmpdirname = tempfile.mkdtemp()
        with open('{}/repofolders.json'.format(tmpdirname), 'w') as _file:
            json.dump(bundle_items, _file)

        self.start_bundle_import(tmpdirname)

        shutil.rmtree(tmpdirname)
        if not self.dry_run:
            transaction.commit()

    def start_bundle_import(self, bundle_path):
        logger.info("\n\nStarting bundle import...\n")
        portal = api.portal.get()
        transmogrifier = Transmogrifier(portal)
        ann = IAnnotations(transmogrifier)
        ann[BUNDLE_PATH_KEY] = bundle_path
        ann[INTERMEDIATE_COMMITS_KEY] = False

        with DisabledLDAP(portal):
            transmogrifier(u'opengever.bundle.oggbundle')

    def move_branches(self, items):
        logger.info("\n\nMoving...\n")
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 1)
            parent = self.guid_to_object(item['new_parent_guid'])
            repo = unrestrictedUuidToObject(item['uid'])
            if not parent or not repo:
                raise Exception('No parent or repo found for {}'.format(item))

            api.content.move(source=repo, target=parent, safe_id=True)
            if not self.dry_run:
                transaction.commit()

    def merge_branches(self, items):
        logger.info("\n\nMerging...\n")
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 1)
            target = self.guid_to_object(item['new_parent_guid'])
            repo = unrestrictedUuidToObject(item['uid'])
            if not target or not repo:
                raise Exception('No target or repo found for {}'.format(item))

            for obj in repo.contentValues():
                api.content.move(source=obj, target=target, safe_id=True)
                self.add_to_reindexing_queue(
                    obj.UID(), ('Title', 'sortable_title', 'reference', 'sortable_reference'),
                    with_children=True)

            deleter = RepositoryDeleter(repo)
            if not deleter.is_deletion_allowed():
                raise Exception('Trying to delete not empty object {}'.format(item))
            deleter.delete()

            if item['uid'] in self.to_reindex:
                self.to_reindex.pop(item['uid'])
            if not self.dry_run:
                transaction.commit()

    def adjust_reference_number_prefix(self, items):
        logger.info("\n\nAdjusting reference number prefix...\n")
        parents = set()
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 5)
            repo = unrestrictedUuidToObject(item['uid'])
            referenceprefix.IReferenceNumberPrefix(repo).reference_number_prefix = item['new_number']
            parents.add(aq_parent(aq_inner(repo)))
            self.add_to_reindexing_queue(
                item['uid'], ('Title', 'sortable_title', 'reference', 'sortable_reference'),
                with_children=True)
        if not self.dry_run:
            transaction.commit()

        self.regenerate_reference_number_mapping(list(parents))

    def regenerate_reference_number_mapping(self, objs):
        logger.info("\n\nRegenerating number mappings...\n")
        for obj in objs:
            ref_adapter = IReferenceNumberPrefix(obj)
            # This purges also the dossier mapping, but we regenerate them
            # below.
            ref_adapter.purge_mappings()

            for child in obj.listFolderContents():
                ref_adapter.set_number(
                    child, number=IReferenceNumber(child).get_local_number())

    def rename(self, items):
        logger.info("\n\nRenaming...\n")
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 1)
            repo = unrestrictedUuidToObject(item['uid'])

            # Rename
            repo.title_de = item['new_title']

            # Adjust id if necessary
            ObjectIDUpdater(repo, FakeOptions()).maybe_update_id()

            # We do not need to reindex path as this seems to already happen
            # recursively
            self.add_to_reindexing_queue(
                item['uid'], ('Title', 'sortable_title'))
            if not self.dry_run:
                transaction.commit()

    def update_description(self, items):
        logger.info("\n\nUpdating descriptions...\n")
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 5)
            repo = unrestrictedUuidToObject(item['uid'])
            if not repo:
                continue

            new_description = item['new_repo_pos'].description
            if repo.description != new_description:
                repo.description = new_description
                self.add_to_reindexing_queue(item['uid'], ('Description',))
        if not self.dry_run:
            transaction.commit()

    def set_permissions(self, items):
        logger.info("\n\nUpdating permissions...\n")
        n_tot = len(items)
        for i, item in enumerate(items):
            log_progress(i, n_tot, 5)
            repo = unrestrictedUuidToObject(item['uid'])
            self._set_permissions_on_object(repo, item['permissions'])
            if not self.dry_run:
                transaction.commit()

    def _set_permissions_on_object(self, obj, permissions):
        """ We set the local roles and block inheritance if needed.
        local_roles are only set if the inheritance is blocked.
        Other conditions should have risen a validation error for the
        excel.
        """
        block_inheritance = permissions['block_inheritance']

        roles_by_principals = defaultdict(list)
        for role_shortname in managed_roles_shortnames:
            role = ROLES_BY_SHORTNAME[role_shortname]
            principals = permissions.get(role_shortname)
            for principal in principals:
                roles_by_principals[principal].append(role)

        obj.__ac_local_roles_block__ = block_inheritance
        manager = RoleAssignmentManager(obj)
        manager.storage.clear_by_cause(ASSIGNMENT_VIA_SHARING)
        for principal, roles in roles_by_principals.items():
            assignment = SharingRoleAssignment(principal, roles)
            manager.storage.add_or_update(
                assignment.principal, assignment.roles, assignment.cause, assignment.reference)

        manager._update_local_roles()

    def reindex(self):
        logger.info("\n\nReindexing...\n")
        n_tot = len(self.to_reindex)
        for i, (uid, idxs) in enumerate(self.to_reindex.items()):
            log_progress(i, n_tot)
            obj = unrestrictedUuidToObject(uid)
            if not obj:
                logger.error("Could not find {} to reindex. Skipping".format(uid))
                continue
            obj.reindexObject(idxs=idxs)
            if obj.portal_type == 'opengever.task.task':
                # make sure that the model is up to date.
                TaskSqlSyncer(obj, None).sync()

    def validate(self):
        """This steps make sure that the repository system has
        been correctly migrated."""
        logger.info("\n\nValidating...\n")
        self.validation_errors = defaultdict(list)
        self.validation_failed = False

        n_tot = len(self.operations_list)
        for i, operation in enumerate(self.operations_list):
            log_progress(i, n_tot)
            # Three possibilities here: position was created, deleted or modified
            if operation['new_position_guid']:
                # new position was created
                obj = self.guid_to_object(operation['new_position_guid'])
            elif operation['uid']:
                obj = unrestrictedUuidToObject(operation['uid'])
                if operation['need_merge']:
                    # position was deleted
                    if obj:
                        logger.error(u"Positions wasn't deleted correctly {}.".format(operation['uid']))
                        self.validation_failed = True
                    continue
            else:
                logger.error(u"Invalid operation {}".format(operation))
                self.validation_failed = True
                continue

            if not obj:
                uid = operation['new_position_guid'] or operation['uid']
                logger.error(u"Could not resolve object {}. Skipping validation.".format(uid))
                self.validation_failed = True
                continue

            # Assert reference number, title and description on the object
            uid = obj.UID()
            new = operation['new_repo_pos']
            self.assertEqual(uid, new.position, obj.get_repository_number().replace('.', ''), 'incorrect number')
            self.assertEqual(uid, new.title, obj.title_de, 'incorrect title')
            self.assertEqual(uid, new.description, obj.description, 'incorrect description')

            # Assert that data in the catalog is consistent with data on the object
            self.checkObjectConsistency(obj)

            if operation['need_creation']:
                # Check that metadata was set correctly
                self.check_metadata(obj, operation)

            # Store some migration information on the object
            IAnnotations(obj)[MIGRATION_KEY] = {
                'old_position': operation['old_repo_pos'].position,
                'new_position': operation['new_repo_pos'].position,
                'old_title': operation['old_repo_pos'].title,
                'new_title': operation['new_repo_pos'].title,
                'old_description': operation['old_repo_pos'].description,
                'new_description': operation['new_repo_pos'].description,
                'new_parent_guid': operation['new_parent_guid'],
                'need_move': operation['need_move'],
                'need_merge': operation['need_merge'],
                'need_creation': operation['need_creation'],
                'permissions': operation['permissions']
            }

        if self.validation_failed:
            raise MigrationValidationError("See log for details")

    def check_metadata(self, obj, operation):
        err_msg = "metadata not set correctly"
        uid = obj.UID()
        metadata = operation["metadata"]
        for field in metadata_fields:
            if field.getName() not in metadata:
                continue
            if object_has_value_for_field(obj, field):
                persisted_value = get_persisted_value_for_field(obj, field)
            else:
                persisted_value = None
            self.checkEqual(uid, persisted_value,
                            metadata.get(field.getName()), err_msg)

    def checkObjectConsistency(self, obj):
        err_msg = "data inconsistency"
        uid = obj.UID()
        brain = uuidToCatalogBrain(uid)
        catalog_data = self.get_catalog_indexdata(obj)

        # reference number obtained through the adapter is generated
        # dynamically, hence it should always be correct.
        # reference number in the catalog and in the metadata should match it.
        refnum = IReferenceNumber(obj).get_number()
        self.checkEqual(uid, refnum, brain.reference, err_msg)
        self.checkEqual(uid, refnum, catalog_data['reference'], err_msg)

        self.checkEqual(uid, brain.Description, obj.Description(), err_msg)

        self.checkEqual(uid, brain.getPath(), obj.absolute_url_path(), err_msg)
        self.checkEqual(uid, catalog_data['path'], obj.absolute_url_path(), err_msg)

        if not obj.portal_type == 'opengever.repository.repositoryfolder':
            return
        self.checkEqual(uid, brain.title_de, obj.get_prefixed_title_de(), err_msg)
        self.checkEqual(uid, brain.title_fr, obj.get_prefixed_title_fr(), err_msg)
        self.checkEqual(uid, catalog_data['sortable_title'], sortable_title(obj)(), err_msg)

    def assertEqual(self, uid, first, second, msg='not equal'):
        """Tests whether first and second are equal as determined by the '=='
        operator. If not, adds error to self.validation_errors, set
        self.validation_failed to true and log the error.
        """
        if not first == second:
            self.validation_errors[uid].append((first, second, msg))
            self.validation_failed = True
            logger.error(u"{}: {} ({}, {})".format(uid, msg, first, second))

    def checkEqual(self, uid, first, second, msg='not equal'):
        """Tests whether first and second are equal as determined by the '=='
        operator. If not, adds error to self.validation_errors, set
        self.validation_failed to true and log the error.
        """
        if not first == second:
            self.validation_errors[uid].append((first, second, msg))
            logger.error(u"{}: {} ({}, {})".format(uid, msg, first, second))

    def get_catalog_indexdata(self, obj):
        """Return the catalog index data for an object as dict.
        """
        rid = self.catalog.getrid('/'.join(obj.getPhysicalPath()))
        return self.catalog.getIndexDataForRID(rid)


class TaskSyncer(object):

    def __init__(self, tasks_to_sync):
        self.tasks_to_sync = tasks_to_sync

    def __call__(self):
        """Syncs all plone tasks with their model
        """
        for uid in self.tasks_to_sync:
            obj = unrestrictedUuidToObject(uid)
            obj.sync()


class FakeOptions(object):
    dry_run = False


def main():
    parser = setup_option_parser()
    parser.add_option(
        '-o', dest='output_directory',
        default='var/migration-{}'.format(MIGRATIOM_TIMESTAMP),
        help='Path to the output directory')
    parser.add_option("-t", "--sync-task", action="store_true",
                      dest="sync_task", default=False)
    parser.add_option("-n", "--dry-run", action="store_true",
                      dest="dryrun", default=False)
    (options, args) = parser.parse_args()

    if not len(args) == 1:
        logger.info("Missing argument, a path to the mapping xlsx")
        sys.exit(1)
    mapping_path = args[0]

    if os.path.isdir(options.output_directory):
        logger.info("Output directory already exists")
        sys.exit(1)

    if not options.output_directory:
        logger.info("Invalid output directory")
        sys.exit(1)

    os.mkdir(options.output_directory)

    if options.dryrun:
        logger.info("Dry run, dooming transaction")
        transaction.doom()

    app = setup_app()
    setup_plone(app, options)

    logger.info('\n\npatching bundle sections...\n')
    PatchCommitSection()()
    PatchReindexContainersSection()()
    PatchReportSection()()
    if not options.sync_task:
        SkipTaskSyncWith()()
    PatchDisableLDAP()()
    SkipDocPropsUpdate()()
    SkipSearchableTextExtraction()()

    logger.info('\n\nstarting analysis...\n')
    analyser = RepositoryExcelAnalyser(mapping_path, options.output_directory)
    analyser.analyse()

    logger.info('\n\nwriting analysis excel...\n')
    analyser.export_to_excel()

    analyser_path = os.path.join(options.output_directory, "analyser.json")
    with open(analyser_path, "w") as outfile:
        analyser_data = {
            'position_mapping': {
                'old_pos_guid': analyser.positions_mapping.old_pos_guid,
                'new_pos_guid': analyser.positions_mapping.new_pos_guid,
                'old_pos_new_guid': analyser.positions_mapping.old_pos_new_guid
            },
            'analysed_rows': analyser.analysed_rows,
        }
        json.dump(analyser_data, outfile, default=vars)

    if not analyser.is_valid:
        logger.info('\n\nInvalid migration excel, aborting...\n')
        return

    migrator = RepositoryMigrator(analyser.analysed_rows, dry_run=options.dryrun)

    logger.info('\n\nstarting migration...\n')
    migrator.run()

    if not options.dryrun:
        logger.info('\n\nCommitting transaction...\n')
        transaction.commit()
        logger.info('Finished migration.')

    migrator_path = os.path.join(options.output_directory, "migrator.json")
    with open(migrator_path, "w") as outfile:
        migrator_data = {
            'operations_list': migrator.operations_list,
            'to_reindex': migrator.to_reindex.keys(),
            'validation_errors': migrator.validation_errors,
        }
        json.dump(migrator_data, outfile, default=vars)

    tasks_to_sync_path = os.path.join(
        options.output_directory, "tasks_to_sync.json")
    with open(tasks_to_sync_path, "w") as outfile:
        json.dump(tuple(tasks_to_sync), outfile)


if __name__ == '__main__':
    main()
