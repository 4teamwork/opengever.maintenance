from collective.transmogrifier.transmogrifier import Transmogrifier
from opengever.bundle.ldap import DisabledLDAP
from opengever.bundle.sections.bundlesource import BUNDLE_PATH_KEY
from opengever.bundle.sections.commit import INTERMEDIATE_COMMITS_KEY
from opengever.dossier.behaviors.dossier import IDossierMarker
from opengever.maintenance.debughelpers import setup_app
from opengever.maintenance.debughelpers import setup_plone
from opengever.repository.interfaces import IRepositoryFolder
from opengever.repository.interfaces import IRepositoryFolderRecords
from opengever.setup.sections.xlssource import xlrd_xls2array
from openpyxl import Workbook
from openpyxl.styles import Font
from plone import api
from zope.annotation import IAnnotations
import argparse
import json
import shutil
import sys
import tempfile


class RepositoryExcelAnalyser(object):

    def __init__(self, mapping_path, analyse_path):
        self.number_changes = {}

        self.diff_xlsx_path = mapping_path
        self.analyse_xlsx_path = analyse_path
        self.analysed_rows = []
        self._reference_repository_mapping = None
        self.catalog = api.portal.get_tool('portal_catalog')

    def analyse(self):
        sheets = xlrd_xls2array(self.diff_xlsx_path)
        if len(sheets) > 1:
            raise Exception('multiple sheets')

        data = sheets[0]['sheet_data']

        # Start on row 16 anything else is header
        for row in data[16:]:
            new_item = {}
            if row[0] in ['', u'l\xf6schen']:
                new_item['position'], new_item['title'], new_item['description'] = None, None, None
            else:
                new_item['position'], new_item['title'] = row[0].split(' ', 1)
                new_item['description'] = row[1]

            if row[2] == '':
                old_item = {'position': None, 'title': None,'description': None}
            else:
                old_item = {'position': str(row[2]), 'title': row[4],
                            'description': row[5]}

            # Remove splitting dots - they're not usefull for comparing etc.
            if old_item['position']:
                old_item['position'] = old_item['position'].replace('.', '')
            if new_item['position']:
                new_item['position'] = new_item['position'].replace('.', '')

            # Ignore empty rows
            if not old_item['position'] and  not new_item['position']:
                continue

            new_number = None
            new_parent_position = None
            parent_of_new_position = None

            needs_creation = not bool(old_item['position'])
            need_number_change, need_move = self.needs_number_change_or_move(new_item, old_item)
            if need_number_change:
                new_number = self.get_new_number(new_item)
            if need_move:
                new_parent_position = self.get_new_parent_position(new_item)
            if needs_creation:
                parent_of_new_position = self.get_parent_of_new_position(new_item)

            analyse = {
                'new_position': parent_of_new_position,
                'new_title': self.get_new_title(new_item, old_item),
                'new_number': new_number,
                'new_parent_position': new_parent_position,
                'old_item': old_item,
                'new_item': new_item,
                'repository_depth_violated': self.is_repository_depth_violated(
                    new_item, old_item),
                'leaf_node_violated': need_move and self.is_leaf_node_principle_violated(
                    new_item, old_item)
            }

            self.analysed_rows.append(analyse)

    def get_new_title(self, new_item, old_item):
        """Returns the new title or none if no rename is necessary."""
        if new_item['title'] != old_item['title']:
            return new_item['title']

        return None

    def get_new_number(self, new_item):
        """Returns latest part of the position - the new referencenumber
        prefix"""
        return new_item['position'][-1]

    def get_new_parent_position(self, new_item):
        """Returns the new parent position"""
        return new_item['position'][:-1]

    def get_parent_of_new_position(self, new_item):
        final_parent_number = new_item['position'][:-1]

        parent_row = [item for item in self.analysed_rows
                      if item['new_item']['position'] == final_parent_number]

        if not parent_row:
            return final_parent_number

        # The parent will be moved to the right position so we need to add
        # the subrepofolder on the "old position"
        return parent_row[0]['old_item']['position']


    def needs_move(self, new_item, old_item):
        if not old_item['position'] or not new_item['position']:
            return False
        parent_old = old_item['position'][:-1]
        parent_new = new_item['position'][:-1]

        if parent_new != parent_old:
            return True

        return False

    def needs_number_change_or_move(self, new_item, old_item):
        """Check if a number change or even a move is necessary
        """
        need_number_change = False
        need_move = False

        if new_item['position'] and old_item['position']:
            if new_item['position'] != old_item['position']:
                need_number_change = True
                self.number_changes[new_item['position']] = old_item['position']

                # check if parent is already changed - so no need to change
                parent_position = new_item['position'][:-1]
                if parent_position in self.number_changes:
                    if self.number_changes[parent_position] == old_item['position'][:-1]:
                        need_number_change = False

                if need_number_change:
                    # check if move is necessary
                    if new_item['position'][:-1] != old_item['position'][:-1]:
                        need_move = True

        return need_number_change, need_move

    def is_repository_depth_violated(self, new_item, old_item):
        max_depth = api.portal.get_registry_record(
            interface=IRepositoryFolderRecords, name='maximum_repository_depth')

        if new_item['position'] and len(new_item['position']) > max_depth:
            return True

        return False

    def is_leaf_node_principle_violated(self, new_item, old_item):
        parent_number = new_item['position'][:-1]
        parent_repo = self.get_repository_reference_mapping().get(parent_number)
        if not parent_repo:
            # Parent does not exist yet, so nothing to worry about it
            return False

        has_dossiers = any([IDossierMarker.providedBy(item) for item in
                            parent_repo.objectValues()])
        return has_dossiers

    def get_repository_reference_mapping(self):
        if not self._reference_repository_mapping:
            repos = [brain.getObject() for brain in
                     self.catalog(object_provides=IRepositoryFolder.__identifier__)]
            self._reference_repository_mapping = {
                repo.get_repository_number(): repo for repo in repos}

        return self._reference_repository_mapping

    def export_to_excel(self):
        workbook = self.prepare_workbook(self.analysed_rows)
        # Save the Workbook-data in to a StringIO
        return workbook.save(filename=self.analyse_xlsx_path)

    def prepare_workbook(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Analyse'

        self.insert_label_row(sheet)
        self.insert_value_rows(sheet, rows)

        return workbook

    def insert_label_row(self, sheet):
        title_font = Font(bold=True)
        labels = [
            # metadata
            'Neu: Position', 'Neu: Titel', 'Neu: Description',
            'Alt: Position', 'Alt: Titel', 'Alt: Description',

            # operations
            'Position Erstellen (Parent Aktenzeichen)',
            'Umbenennung (Neuer Titel)',
            'Nummer Anpassung (Neuer `Praefix`)',
            'Verschiebung (Aktenzeichen neues Parent)',

            # rule violations
            'Verletzt Max. Tiefe',
            'Verletzt Leafnode Prinzip'
        ]

        for i, label in enumerate(labels, 1):
            cell = sheet.cell(row=1 + 1, column=i)
            cell.value = label
            cell.font = title_font

    def insert_value_rows(self, sheet, rows):
        for row, data in enumerate(rows, 2):
            values = [
                data['new_item']['position'],
                data['new_item']['title'],
                data['new_item']['description'],
                data['old_item']['position'],
                data['old_item']['title'],
                data['old_item']['description'],
                data['new_position'],
                data['new_title'],
                data['new_number'],
                data['new_parent_position'],
                'x' if data['repository_depth_violated'] else '',
                'x' if data['leaf_node_violated'] else '',
            ]

            for column, attr in enumerate(values, 1):
                cell = sheet.cell(row=1 + row, column=column)
                cell.value = attr


class RepositoryMigrator(object):

    def __init__(self, operations_list):
        self.operations_list = operations_list

    def run(self):
        self.create_repository_folders(self.items_to_create())

    def items_to_create(self):
        return [item for item in self.operations_list if item['new_position']]

    def create_repository_folders(self, items):
        """Add repository folders - by using the ogg.bundle import. """

        bundle_items = []
        for item in items:
            # Bundle expect the format [[repository], [dossier]]
            parent_reference = [[int(x) for x in list(item['new_position'])]]
            from uuid import uuid4
            bundle_items.append(
                {'guid': uuid4().hex[:8],
                 'description': item['new_item']['description'],
                 'parent_reference': parent_reference,
                 'reference_number_prefix': item['new_item']['position'][-1],
                 'review_state': 'repositoryfolder-state-active',
                 'title_de': item['new_item']['title']})

        tmpdirname = tempfile.mkdtemp()
        with open('{}/repofolders.json'.format(tmpdirname), 'w') as _file:
            json.dump(bundle_items, _file)

        self.start_bundle_import(tmpdirname)

        shutil.rmtree(tmpdirname)

    def start_bundle_import(self, bundle_path):
        portal = api.portal.get()
        transmogrifier = Transmogrifier(portal)
        ann = IAnnotations(transmogrifier)
        ann[BUNDLE_PATH_KEY] = bundle_path
        ann[INTERMEDIATE_COMMITS_KEY] = False

        with DisabledLDAP(portal):
            transmogrifier(u'opengever.bundle.oggbundle')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', dest='site_root', default=None,
                        help='Absolute path to the Plone site')
    parser.add_argument('-m', dest='mapping', default=None,
                        help='Path to the mapping xlsx')
    parser.add_argument('-o', dest='output', default=None,
                        help='Path to the output xlsx')
    options = parser.parse_args(sys.argv[3:])
    app = setup_app()

    setup_plone(app, options)

    analyser = RepositoryExcelAnalyser(options.mapping, options.output)
    analyser.analyse()
    analyser.export_to_excel()


if __name__ == '__main__':
    main()
