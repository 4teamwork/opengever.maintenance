# -*- coding: utf-8 -*-
from opengever.base.interfaces import IReferenceNumber
from opengever.base.behaviors.classification import IClassification
from opengever.base.behaviors.lifecycle import ILifeCycle
from openpyxl import Workbook
from plone import api
from zope.i18n import translate


class OrdnungsSystemToXlsx(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.set_xlsx_headers()

    def set_xlsx_headers(self):
        self.ws.title = 'Ordnungssystem'
        self.ws.append(
            ['Ordnungs-positions-nummer',
             'Titel der Ordnungsposition',
             'Titel der ordnungsposition (französisch)',
             'Beschreibung (optional)',
             'Klassifikation',
             'Datenschutzstufe',
             'Aufbewahrung in Verwaltung',
             'Kommentar zur Aufbewahrungsdauer (optional)',
             'Archivwürdigkeit',
             'Kommentar zur Archivwürdigkeit (optional)',
             'Archivische Schutzfrist',
             'Gültig ab',
             'Gültig bis'])
        self.ws.append([])

    def create_ordnungssystem_xlsx(self):
        portal = api.portal.get()
        ordnungssystem = portal.ordnungssystem

        self.ws['B3'] = 'Ordnungssystem'
        self.ws['L3'] = ordnungssystem.valid_from
        self.ws['M3'] = ordnungssystem.valid_until

        for node in ordnungssystem.getChildNodes():
            self.get_nodes(node)

        self.wb.save('ordnungssystem.xlsx')

    def get_nodes(self, node):
        if node.__class__.__name__ in ['RepositoryRoot', 'RepositoryFolder']:
            reference_number = IReferenceNumber(node).get_repository_number()

            self.write_node_to_xlsx(reference_number, node)

            print 'added ' + reference_number + ' ' + node.id

            if node.hasChildNodes():
                for child in node.getChildNodes():
                    self.get_nodes(child)

    def write_node_to_xlsx(self, reference_number, node):
        privacy_layer = IClassification(node).privacy_layer
        life_cycle = ILifeCycle(node)

        archival_value = life_cycle.annotations['opengever.base.behaviors.lifecycle.ILifeCycle.archival_value']
        archival_value_annotation = life_cycle.annotations['opengever.base.behaviors.lifecycle.ILifeCycle.archival_value_annotation']
        custody_period = life_cycle.annotations['opengever.base.behaviors.lifecycle.ILifeCycle.custody_period']
        retention_period = life_cycle.annotations['opengever.base.behaviors.lifecycle.ILifeCycle.retention_period']
        retention_period_annotation = life_cycle.annotations['opengever.base.behaviors.lifecycle.ILifeCycle.retention_period_annotation']

        archival_value_de = translate(archival_value, domain='opengever.base', target_language='de')
        classification_de = translate(node.classification, domain='opengever.base', target_language='de')

        if privacy_layer == u'privacy_layer_no':
            privacy = 'Enthält schützenswerte Personendaten'
        else:
            privacy = 'Keine Datenschutzstufe'

        self.ws.append([
            reference_number,
            node.title_de,
            node.title_fr,
            node.description,
            classification_de,
            privacy,
            retention_period,
            retention_period_annotation,
            archival_value_de,
            archival_value_annotation,
            custody_period,
            node.valid_from,
            node.valid_until])


if __name__ == '__main__':
    ordnungssystem_to_xlsx = OrdnungsSystemToXlsx()
    ordnungssystem_to_xlsx.create_ordnungssystem_xlsx()
    print 'Data has been saved to ordnungssystem.xlsx'
